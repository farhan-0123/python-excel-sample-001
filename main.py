from openpyxl import load_workbook, Workbook

# Global Constants
SALES_XLSX_PATH = "Input Excel Files\\faker-sales.xlsx"
TRANSACTIONS_XLSX_PATH = "Input Excel Files\\faker-credit-card.xlsx"
SALES_RECONCILED_PATH = "Output Excel Files\\Sales_reconciled.xlsx"

def replace_old_credit_card_with_new(sales_xlsx, transactions_xlsx):
    # Open a new workbook for output file 
    sales_reconciled = Workbook()

    # Open currenty active sheet in the new file
    sales_reconciled_sheet = sales_reconciled.active
    
    # Open currently active sheet in both input files
    sales_sheet = sales_xlsx.active
    transactions_sheet = transactions_xlsx.active

    # Add headers to the output file
    sales_reconciled_sheet.append(["Name", "New Credit Card Number", "Transaction Date"])

    # Find matching old credit card numbers and replace them with new numbers
    for sale_row in sales_sheet.iter_rows(values_only=True):
        for transaction_row in transactions_sheet.iter_rows(values_only=True):
            if sale_row[1] == transaction_row[1]:
                sales_reconciled_sheet.append([sale_row[0], transaction_row[2], sale_row[2]])
    
    # Save the output file
    sales_reconciled.save(SALES_RECONCILED_PATH)

def main():
    # Load the xlsx files from the Input Folder
    sales_xlsx = load_workbook(SALES_XLSX_PATH)
    transactions_xlsx = load_workbook(TRANSACTIONS_XLSX_PATH)
    
    # Run the script
    replace_old_credit_card_with_new(sales_xlsx, transactions_xlsx)

    pass

# Only run when this is the entry point
if __name__ == "__main__":
    main()